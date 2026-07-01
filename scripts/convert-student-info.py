#!/usr/bin/env python3

import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


HEADER_KEYS = {
    "学号": "studentId",
    "姓名": "name",
    "性别": "gender",
    "出生日期": "birthDate",
    "年龄": "age",
    "年级": "grade",
    "专业": "major",
    "学院": "college",
    "导师编号": "supervisorId",
    "研究方向": "researchArea",
    "已发表B类及以上论文数量": "paperCount",
    "课题参与情况": "projectParticipation",
    "外语水平": "languageLevel",
    "信息核验状态": "verificationStatus",
    "住宿状态": "residenceStatus",
    "培养方式": "trainingMode",
    "奖助情况": "fundingStatus",
    "备注": "notes",
}

EXCLUDED_HEADERS = {
    "已发表B类及以上论文数量",
    "信息核验状态",
    "住宿状态",
    "培养方式",
    "备注",
}


def normalize_value(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return value if value is not None else ""


def convert_workbook(source: Path):
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(row_iterator)]
    keys = [HEADER_KEYS.get(header, f"field{index + 1}") for index, header in enumerate(headers)]
    selected_columns = [
        (index, key, header)
        for index, (key, header) in enumerate(zip(keys, headers))
        if header not in EXCLUDED_HEADERS
    ]
    columns = [
        {"key": key, "label": header}
        for _, key, header in selected_columns
    ]
    rows = []

    for values in row_iterator:
        row = {
            key: normalize_value(values[index])
            for index, key, _ in selected_columns
        }
        if any(value != "" for value in row.values()):
            rows.append(row)

    return {
        "sheet": worksheet.title,
        "columns": columns,
        "rows": rows,
    }


def main():
    parser = argparse.ArgumentParser(description="Convert the ARG student workbook to browser JSON.")
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()

    payload = convert_workbook(args.source)
    args.destination.parent.mkdir(parents=True, exist_ok=True)
    args.destination.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"converted {len(payload['rows'])} rows and "
        f"{len(payload['columns'])} columns from {payload['sheet']}"
    )


if __name__ == "__main__":
    main()
